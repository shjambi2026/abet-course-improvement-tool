from datetime import datetime
from io import BytesIO
import json
import os
import streamlit as st

from services.report_service import (
    generate_coordination_mom,
    generate_assessment_cover_sheet,
    generate_eos_report,
)

try:
    from services.ai_service import ask_ai_json
except Exception:
    ask_ai_json = None

try:
    from services.attainment_service import calculate_so_attainment_from_raw_workbook
except Exception:
    calculate_so_attainment_from_raw_workbook = None


# --------------------------------------------------
# Validate Session
# --------------------------------------------------

if "instructor_profile" not in st.session_state:
    st.warning("Please create your instructor profile first.")
    st.page_link("pages/1_Instructor_Profile.py", label="Go to Instructor Profile", icon="👤")
    st.stop()

if "current_course" not in st.session_state:
    st.warning("Please load a course in Course Workspace first.")
    st.page_link("pages/2_Course_Workspace.py", label="Go to Course Workspace", icon="📚")
    st.stop()

if "course_structure" not in st.session_state:
    st.warning("Please load a course with an articulation matrix in Course Workspace first.")
    st.page_link("pages/2_Course_Workspace.py", label="Go to Course Workspace", icon="📚")
    st.stop()


# --------------------------------------------------
# Load Data
# --------------------------------------------------

course = st.session_state["current_course"]["profile"]
course_structure = st.session_state["course_structure"]

assessments = course_structure.get(
    "assessment_instances",
    course_structure.get("assessments", []),
)


# --------------------------------------------------
# Helpers
# --------------------------------------------------

def clean_filename_text(value):
    return (
        str(value or "")
        .replace("/", "-")
        .replace("\\", "-")
        .replace(" ", "")
        .replace(":", "")
        .replace("•", "-")
    )


def safe_file_part(value):
    return (
        str(value or "")
        .replace("/", "-")
        .replace("\\", "-")
        .replace(" ", "_")
        .replace(":", "")
        .replace("•", "-")
    )


def download_markdown(label, text, filename):
    st.download_button(
        label=label,
        data=text.encode("utf-8"),
        file_name=filename,
        mime="text/markdown",
    )


def assessment_label(assessment):
    week = assessment.get("week", "")
    name = assessment.get("name", "Assessment")
    so_text = format_so_distribution(assessment)
    return f"Week {week} • {name} • {so_text}"


def show_preview_and_download(state_key, label, filename):
    if st.session_state.get(state_key):
        st.markdown("### Preview")
        st.text_area(
            label,
            value=st.session_state[state_key],
            height=450,
        )

        download_markdown(
            f"📥 Download {label}",
            st.session_state[state_key],
            filename,
        )


def get_assessment_id(assessment):
    return assessment.get("id") or clean_filename_text(assessment.get("name", "Assessment"))


def sort_so_ids(values):
    def key(value):
        try:
            return int(str(value).replace("SO", ""))
        except Exception:
            return 9999

    return sorted(values or [], key=key)


def format_so_distribution(assessment):
    distribution = assessment.get("so_distribution", {})

    if distribution:
        items = sorted(
            distribution.items(),
            key=lambda item: int(str(item[0]).replace("SO", "")),
        )
        return ", ".join(f"{so} ({value:g}%)" for so, value in items)

    sos = assessment.get("assessed_sos", assessment.get("sos", []))
    return ", ".join(sort_so_ids(sos)) if sos else "—"


def course_upload_folder(course):
    course_id = course.get("course_id") or clean_filename_text(
        f"{course.get('course_code', 'Course')}_{course.get('academic_year', '')}_{course.get('semester', '')}"
    )
    return os.path.join("so_attainment_uploads", safe_file_part(course_id))


def saved_marks_paths(course, assessment_id):
    folder = course_upload_folder(course)
    safe_id = safe_file_part(assessment_id)
    return (
        os.path.join(folder, f"{safe_id}.xlsx"),
        os.path.join(folder, f"{safe_id}.json"),
    )


def extract_section_counts_from_workbook(uploaded_bytes):
    """
    Read saved SO marks workbook and count students by section.
    Expected columns: Program | Section | Student ID | SO...
    """

    try:
        import openpyxl

        wb = openpyxl.load_workbook(BytesIO(uploaded_bytes), data_only=True)
        ws = wb["grades"] if "grades" in wb.sheetnames else wb.active

        header_row = None

        for row in range(1, ws.max_row + 1):
            values = [
                str(ws.cell(row=row, column=col).value or "").strip().lower()
                for col in range(1, min(ws.max_column, 8) + 1)
            ]

            if "program" in values and "section" in values and "student id" in values:
                header_row = row
                break

        if not header_row:
            return {}

        counts = {}

        for row in range(header_row + 2, ws.max_row + 1):
            section = ws.cell(row=row, column=2).value
            student_id = ws.cell(row=row, column=3).value

            if not student_id:
                continue

            section = str(section or "Unspecified").strip()
            counts[section] = counts.get(section, 0) + 1

        return counts

    except Exception:
        return {}


def load_saved_so_attainment(course, so_assessments):
    results = {}

    if calculate_so_attainment_from_raw_workbook is None:
        return results

    for assessment in so_assessments:
        aid = get_assessment_id(assessment)
        xlsx_path, meta_path = saved_marks_paths(course, aid)

        if not os.path.exists(xlsx_path):
            continue

        try:
            with open(xlsx_path, "rb") as f:
                uploaded_bytes = f.read()

            metadata = {}

            if os.path.exists(meta_path):
                with open(meta_path, "r", encoding="utf-8") as f:
                    metadata = json.load(f)

            attainment = calculate_so_attainment_from_raw_workbook(BytesIO(uploaded_bytes))
            section_counts = extract_section_counts_from_workbook(uploaded_bytes)

            results[aid] = {
                "assessment": assessment,
                "attainment": attainment,
                "section_counts": section_counts,
                "uploaded_filename": metadata.get("uploaded_filename", f"{aid}.xlsx"),
                "uploaded_bytes": uploaded_bytes,
            }

        except Exception:
            continue

    return results


def calculate_combined_so_attainment(saved_results):
    weighted_totals = {}
    weight_totals = {}
    students_assessed = {}
    students_attained = {}

    for record in saved_results.values():
        assessment = record.get("assessment", {})
        attainment = record.get("attainment", {})
        summary = attainment.get("summary", [])

        course_weight = assessment.get("course_weight", assessment.get("weight", 0))
        so_distribution = assessment.get("so_distribution", {})

        summary_by_so = {item.get("so"): item for item in summary}

        for so, distribution_percent in so_distribution.items():
            if so not in summary_by_so:
                continue

            item = summary_by_so[so]
            weight = course_weight * (distribution_percent / 100)

            weighted_totals[so] = weighted_totals.get(so, 0) + (
                item.get("attainment_percent", 0) * weight
            )
            weight_totals[so] = weight_totals.get(so, 0) + weight

            students_assessed[so] = max(
                students_assessed.get(so, 0),
                item.get("students_assessed", 0),
            )
            students_attained[so] = max(
                students_attained.get(so, 0),
                item.get("students_attained", 0),
            )

    combined_summary = []

    for so in sorted(weighted_totals.keys(), key=lambda x: int(str(x).replace("SO", ""))):
        total_weight = weight_totals.get(so, 0)
        attainment_percent = round(weighted_totals[so] / total_weight, 2) if total_weight else 0

        combined_summary.append(
            {
                "so": so,
                "combined_weight": round(total_weight, 2),
                "students_assessed": students_assessed.get(so, "—"),
                "students_attained": students_attained.get(so, "—"),
                "attainment_percent": attainment_percent,
                "target_met": "Yes" if attainment_percent >= 65 else "No",
            }
        )

    return {
        "summary": combined_summary,
        "source": "Saved SO marks workbooks",
    }


def get_default_total_marks(course, assessment):
    aid = get_assessment_id(assessment)
    xlsx_path, _ = saved_marks_paths(course, aid)

    if not os.path.exists(xlsx_path):
        return 30.0

    try:
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        ws = wb["grades"] if "grades" in wb.sheetnames else wb.active

        header_row = None

        for row in range(1, ws.max_row + 1):
            values = [
                str(ws.cell(row=row, column=col).value or "").strip().lower()
                for col in range(1, min(ws.max_column, 8) + 1)
            ]

            if "student id" in values:
                header_row = row
                break

        if not header_row:
            return 30.0

        max_row = header_row + 1
        total = 0.0

        for col in range(4, ws.max_column + 1):
            value = ws.cell(row=max_row, column=col).value
            if isinstance(value, (int, float)):
                total += float(value)

        return total if total > 0 else 30.0

    except Exception:
        return 30.0


def get_default_number_of_students(saved_results):
    for key in ["number_of_students", "students", "num_students"]:
        if course.get(key):
            return str(course.get(key))

    max_students = 0

    for record in saved_results.values():
        for item in record.get("attainment", {}).get("summary", []):
            value = item.get("students_assessed", 0)
            if isinstance(value, (int, float)):
                max_students = max(max_students, int(value))

    return str(max_students) if max_students else ""


def render_markdown_table(headers, rows):
    header = "| " + " | ".join(headers) + " |\n"
    separator = "|" + "|".join(["---" for _ in headers]) + "|\n"
    body = "\n".join("| " + " | ".join(str(value) for value in row) + " |" for row in rows)
    st.markdown(header + separator + body)


def render_section_and_attainment_intro(saved_results, combined_so_attainment):
    st.markdown("### Section and Student Summary")

    section_totals = {}

    for record in saved_results.values():
        for section, count in record.get("section_counts", {}).items():
            section_totals[section] = max(section_totals.get(section, 0), count)

    if section_totals:
        render_markdown_table(
            ["Section", "Students"],
            [[section, count] for section, count in sorted(section_totals.items())],
        )
    else:
        st.info("Section counts were not detected from the saved marks workbooks.")

    if combined_so_attainment and combined_so_attainment.get("summary"):
        st.markdown("### Current SO Attainment Results")
        render_markdown_table(
            ["SO", "Students Assessed", "Students Attained", "Attainment", "Target Met"],
            [
                [
                    item.get("so", ""),
                    item.get("students_assessed", "—"),
                    item.get("students_attained", "—"),
                    f"{item.get('attainment_percent', 0)}%",
                    item.get("target_met", "No"),
                ]
                for item in combined_so_attainment.get("summary", [])
            ],
        )
    else:
        st.info(
            "No saved SO attainment results were found yet. You can still generate a draft EOS report, "
            "but the SO Attainment Summary section will be incomplete."
        )


def render_assessment_comparison(saved_results):
    st.markdown("### Assessment-Level Comparison")

    if not saved_results:
        st.info("No assessment-level results are available yet.")
        return

    rows = []

    for record in saved_results.values():
        assessment = record.get("assessment", {})
        summary = record.get("attainment", {}).get("summary", [])

        so_text = ", ".join(
            f"{item.get('so', '')}: {item.get('attainment_percent', 0)}%"
            for item in summary
        )

        rows.append(
            [
                assessment.get("name", ""),
                format_so_distribution(assessment),
                so_text or "—",
            ]
        )

    render_markdown_table(
        ["Assessment", "SO Mapping (Distribution)", "Attainment"],
        rows,
    )


def build_assessment_based_evaluation_text(
    saved_results,
    combined_so_attainment,
    section_text,
    exam_feedback,
    project_feedback,
    comparison_feedback,
    instructor_notes,
):
    parts = []

    parts.append("The SO attainment results were reviewed based on the saved assessment-level evidence.")

    if section_text:
        parts.append(section_text)

    if combined_so_attainment and combined_so_attainment.get("summary"):
        summary_sentences = []

        for item in combined_so_attainment.get("summary", []):
            summary_sentences.append(
                f"{item.get('so', '')} achieved {item.get('attainment_percent', 0)}% attainment "
                f"with target met: {item.get('target_met', 'No')}."
            )

        parts.append(" ".join(summary_sentences))

    if exam_feedback and exam_feedback != "Not applicable":
        parts.append(f"For the exam-based assessment, {exam_feedback.lower()}.")

    if project_feedback and project_feedback != "Not applicable":
        parts.append(f"For the project-based assessment, {project_feedback.lower()}.")

    if comparison_feedback and comparison_feedback != "Not applicable":
        parts.append(f"When comparing the SO-assessing assessments, {comparison_feedback.lower()}.")

    if instructor_notes:
        parts.append(instructor_notes)

    return "\n\n".join(parts)


def build_subjective_evaluation_text(engagement, readiness, participation, instructor_notes):
    parts = []

    if engagement and engagement != "Not specified":
        parts.append(f"Student engagement was {engagement.lower()}.")

    if readiness and readiness != "Not specified":
        parts.append(f"Student readiness and technical preparation were {readiness.lower()}.")

    if participation and participation != "Not specified":
        parts.append(f"Student participation and interaction were {participation.lower()}.")

    if instructor_notes:
        parts.append(instructor_notes)

    return "\n\n".join(parts) if parts else ""


def build_course_improvement_text(case_studies, hands_on, ai_support, instructor_notes):
    improvements = []

    if case_studies:
        improvements.append("Incorporate additional real-world case studies to strengthen the connection between course concepts and industry practice.")

    if hands_on:
        improvements.append("Expand hands-on modeling exercises to provide students with more opportunities to apply analysis and architectural design techniques.")

    if ai_support:
        improvements.append("Explore effective ways to leverage AI-supported learning activities while maintaining critical thinking and independent problem-solving.")

    if instructor_notes:
        improvements.append(instructor_notes)

    return "\n".join(f"{i}. {item}" for i, item in enumerate(improvements, start=1))


def suggest_eos_text(course, course_structure, saved_results, combined_so_attainment):
    if ask_ai_json is None:
        return {
            "assessment_based_evaluation": "",
            "subjective_evaluation": "",
            "course_improvement": "",
        }

    prompt = f"""
You are helping an instructor draft an End-of-Semester Report.

Use the course data, section/student counts, and SO attainment results to suggest concise text for:
1. Assessment-Based Evaluation
2. Subjective Evaluation
3. Course Improvement

The Assessment-Based Evaluation should:
- mention section/student information if available,
- summarize SO attainment results,
- compare exam and project assessment evidence if both exist,
- avoid inventing student-level data.

Course:
{course}

Course Structure:
{course_structure}

Saved Assessment Results:
{saved_results}

Combined SO Attainment:
{combined_so_attainment}

Return only valid JSON:
{{
  "assessment_based_evaluation": "text",
  "subjective_evaluation": "text",
  "course_improvement": "text"
}}
"""

    return ask_ai_json(prompt)




def craft_eos_text_with_ai(
    course,
    course_structure,
    saved_results,
    combined_so_attainment,
    section_text,
    exam_feedback,
    project_feedback,
    comparison_feedback,
    assessment_notes,
    engagement,
    readiness,
    participation,
    subjective_notes,
    improvement_flags,
    improvement_notes,
):
    fallback_assessment = build_assessment_based_evaluation_text(
        saved_results=saved_results,
        combined_so_attainment=combined_so_attainment,
        section_text=section_text,
        exam_feedback=exam_feedback,
        project_feedback=project_feedback,
        comparison_feedback=comparison_feedback,
        instructor_notes=assessment_notes,
    )

    fallback_subjective = build_subjective_evaluation_text(
        engagement=engagement,
        readiness=readiness,
        participation=participation,
        instructor_notes=subjective_notes,
    )

    fallback_improvement = build_course_improvement_text(
        case_studies=improvement_flags.get("case_studies", False),
        hands_on=improvement_flags.get("hands_on", False),
        ai_support=improvement_flags.get("ai_support", False),
        instructor_notes=improvement_notes,
    )

    if ask_ai_json is None:
        return {
            "assessment_based_evaluation": fallback_assessment,
            "subjective_evaluation": fallback_subjective,
            "course_improvement": fallback_improvement,
        }

    prompt = f"""
You are helping an instructor write the final narrative sections of an End-of-Semester Report.

Use ALL instructor-selected insights and any additional instructor notes. 
Write complete, polished, smoothly connected, professional text.
Do not invent student-level data. Use only aggregate attainment results and instructor-provided insights.

Course:
{course}

Course Structure:
{course_structure}

Saved Assessment Results:
{saved_results}

Combined SO Attainment:
{combined_so_attainment}

Instructor Inputs:
Section/student context: {section_text}
Exam feedback: {exam_feedback}
Project feedback: {project_feedback}
Comparison feedback: {comparison_feedback}
Assessment notes: {assessment_notes}
Student engagement: {engagement}
Student readiness: {readiness}
Participation and interaction: {participation}
Subjective notes: {subjective_notes}
Improvement options: {improvement_flags}
Additional improvement notes: {improvement_notes}

Return only valid JSON:
{{
  "assessment_based_evaluation": "one or two complete paragraphs",
  "subjective_evaluation": "one complete paragraph",
  "course_improvement": "a concise numbered list"
}}
"""

    try:
        result = ask_ai_json(prompt)

        return {
            "assessment_based_evaluation": result.get("assessment_based_evaluation", fallback_assessment),
            "subjective_evaluation": result.get("subjective_evaluation", fallback_subjective),
            "course_improvement": result.get("course_improvement", fallback_improvement),
        }
    except Exception:
        return {
            "assessment_based_evaluation": fallback_assessment,
            "subjective_evaluation": fallback_subjective,
            "course_improvement": fallback_improvement,
        }


# --------------------------------------------------
# Prepare saved SO attainment
# --------------------------------------------------

so_assessments = [
    assessment for assessment in assessments
    if assessment.get("assessing_so", False)
]

saved_attainment_results = st.session_state.get("so_attainment_results", {})
persisted_attainment_results = load_saved_so_attainment(course, so_assessments)

if persisted_attainment_results:
    saved_attainment_results.update(persisted_attainment_results)
    st.session_state["so_attainment_results"] = saved_attainment_results

combined_so_attainment = st.session_state.get(
    "combined_so_attainment",
    st.session_state.get("course_evidence", {}).get("combined_so_attainment", {}),
)

if saved_attainment_results:
    combined_so_attainment = calculate_combined_so_attainment(saved_attainment_results)
    st.session_state["combined_so_attainment"] = combined_so_attainment
    st.session_state.setdefault("course_evidence", {})
    st.session_state["course_evidence"]["combined_so_attainment"] = combined_so_attainment


# --------------------------------------------------
# Page Header
# --------------------------------------------------

st.title("📄 Report Generator")

st.write(
    "Select one of the report areas below to generate the required course documents "
    "in the accreditation workflow order."
)

st.markdown(f"## {course.get('course_code', '')} – {course.get('course_name', '')}")

course_code = clean_filename_text(course.get("course_code", "Course"))
academic_year = clean_filename_text(course.get("academic_year", ""))
semester = clean_filename_text(course.get("semester", ""))
today = datetime.now().strftime("%Y-%m-%d")


# --------------------------------------------------
# Report Selection Areas
# --------------------------------------------------

st.markdown("---")
st.markdown("## Select Report")


# ==================================================
# 1. Coordination MOM
# ==================================================

with st.expander("1️⃣ Coordination MOM", expanded=False):
    st.write("Generate course coordination meeting minutes using the department MOM structure.")

    c1, c2 = st.columns(2)

    with c1:
        meeting_date = st.date_input("Meeting Date")
        meeting_no = st.text_input("Meeting #", value="1")
        meeting_time = st.text_input("Meeting Time", placeholder="Example: 2:00 PM")
        attendees = st.text_area(
            "Attendance / Names",
            placeholder="Enter attendees, one per line.",
            height=120,
        )

    with c2:
        topics_discussed = st.text_area(
            "Topics Discussed / Actions Taken / Points Agreed",
            placeholder="Enter each topic, action, or point agreed on a separate line.",
            height=280,
        )

        next_meeting = st.text_input(
            "Next Meeting",
            placeholder="Example: Will be arranged / End of semester",
        )

    if st.button("Generate Coordination MOM", type="primary"):
        mom_text = generate_coordination_mom(
            course=course,
            course_structure=course_structure,
            meeting_no=meeting_no,
            meeting_date=str(meeting_date),
            meeting_time=meeting_time,
            attendees=attendees,
            topics_discussed=topics_discussed,
            prepared_by=course.get("coordinator", course.get("instructor", "")),
            next_meeting=next_meeting,
        )

        st.session_state["coordination_mom_report"] = mom_text

    filename = f"Coordination_MOM_{course_code}_{academic_year}_{semester}_{today}.md"
    show_preview_and_download("coordination_mom_report", "Coordination MOM", filename)


# ==================================================
# 2. Assessment Coversheet
# ==================================================

with st.expander("2️⃣ Assessment Coversheet", expanded=False):
    st.write("Generate a coversheet for SO-assessing assessments only.")

    if not so_assessments:
        st.warning("No SO-assessing assessments were detected from the course structure.")
    else:
        selected_assessment = st.selectbox(
            "Assessment",
            so_assessments,
            format_func=assessment_label,
            key="cover_sheet_assessment",
        )

        default_marks = get_default_total_marks(course, selected_assessment)

        c1, c2 = st.columns(2)

        with c1:
            assessment_date = st.date_input("Assessment Date", key="cover_sheet_date")

            total_marks = st.number_input(
                "Total Marks",
                min_value=1.0,
                value=float(default_marks),
                step=1.0,
            )

        with c2:
            duration = st.text_input("Duration", placeholder="Example: 90 minutes")

            section = st.text_input(
                "Section",
                value=course.get("sections", course.get("section", "")),
            )

        notes = st.text_area("Notes", placeholder="Optional notes.", height=100)

        if st.button("Generate Assessment Coversheet", type="primary"):
            cover_sheet = generate_assessment_cover_sheet(
                course=course,
                assessment=selected_assessment,
                assessment_date=str(assessment_date),
                duration=duration,
                total_marks=total_marks,
                section=section,
                program="IS",
                notes=notes,
            )

            st.session_state["assessment_cover_sheet_report"] = cover_sheet
            st.session_state["assessment_cover_sheet_name"] = selected_assessment.get("name", "Assessment")

        assessment_name = clean_filename_text(st.session_state.get("assessment_cover_sheet_name", "Assessment"))

        filename = f"Assessment_Cover_Sheet_{assessment_name}_{course_code}_{academic_year}_{semester}_{today}.md"
        show_preview_and_download("assessment_cover_sheet_report", "Assessment Cover Sheet", filename)


# ==================================================
# 3. End-of-Semester Report
# ==================================================

with st.expander("3️⃣ End-of-Semester Report (EOS)", expanded=False):
    st.write(
        "Generate the EOS report using section/student information, SO attainment results, "
        "instructor insights, and AI-crafted final write-ups."
    )

    render_section_and_attainment_intro(saved_attainment_results, combined_so_attainment)
    render_assessment_comparison(saved_attainment_results)

    default_students = get_default_number_of_students(saved_attainment_results)
    default_sections = course.get("sections", course.get("section", ""))

    if default_sections and default_students:
        default_section_context = (
            f"The course was offered to section(s) {default_sections} with "
            f"{default_students} students."
        )
    elif default_sections:
        default_section_context = f"The course was offered to section(s) {default_sections}."
    else:
        default_section_context = ""

    number_of_students = st.text_input(
        "Number of Students",
        value=default_students,
        placeholder="Example: 11",
    )

    st.markdown("### 1. AI First Draft")

    if st.button("✨ Suggest EOS Text with AI", type="primary"):
        with st.spinner("Generating complete suggested EOS text..."):
            try:
                suggestions = suggest_eos_text(
                    course,
                    course_structure,
                    saved_attainment_results,
                    combined_so_attainment,
                )
                st.session_state["eos_assessment_based_evaluation"] = suggestions.get("assessment_based_evaluation", "")
                st.session_state["eos_subjective_evaluation"] = suggestions.get("subjective_evaluation", "")
                st.session_state["eos_course_improvement"] = suggestions.get("course_improvement", "")
                st.success("AI first draft generated. You can add instructor insights below and merge them into the draft.")
            except Exception as e:
                st.error("Could not generate EOS suggestions.")
                st.caption(str(e))

    st.markdown("### 2. Add Instructor Insights")

    section_text = st.text_area(
        "Section and student context",
        value=st.session_state.get("eos_section_context", default_section_context),
        placeholder="Example: The course was offered to section DAR with 11 students.",
        height=80,
    )
    st.session_state["eos_section_context"] = section_text

    st.markdown("#### Assessment-Based Evaluation Insights")

    c1, c2 = st.columns(2)

    with c1:
        exam_feedback = st.selectbox(
            "Exam-based assessment feedback",
            [
                "Not applicable",
                "Students performed strongly in the exam, indicating good understanding of the assessed concepts",
                "Students showed acceptable exam performance but some difficulty with applied or analytical questions",
                "Students struggled in the exam, suggesting a need for additional review and practice",
            ],
        )

        project_feedback = st.selectbox(
            "Project-based assessment feedback",
            [
                "Not applicable",
                "Project performance was strong and demonstrated students' ability to apply course concepts",
                "Project performance was acceptable but students needed more guidance in applying concepts",
                "Project performance showed gaps in practical application and should be supported with additional scaffolding",
            ],
        )

    with c2:
        comparison_feedback = st.selectbox(
            "Comparison between SO assessments",
            [
                "Not applicable",
                "students performed consistently across the SO-assessing activities",
                "project performance was stronger than exam performance, suggesting better learning through applied work",
                "exam performance was stronger than project performance, suggesting a need to improve project guidance and milestones",
                "performance varied across assessments, suggesting that students need more balanced preparation across assessment types",
            ],
        )

        assessment_notes = st.text_area(
            "Additional assessment-based evaluation notes",
            placeholder="Optional: Add your own interpretation of the attainment results.",
            height=90,
        )

    st.markdown("#### Subjective Evaluation Insights")

    c3, c4 = st.columns(2)

    with c3:
        engagement = st.selectbox(
            "Student engagement",
            [
                "Not specified",
                "high and students were interactive and motivated",
                "good overall with steady participation",
                "mixed, with some students requiring more encouragement and follow-up",                
            ],
        )

        readiness = st.selectbox(
            "Student readiness",
            [
                "Not specified",
                "strong and students were able to handle the course expectations",
                "acceptable, although some students needed additional support",
                "varied, with noticeable differences in students' prior preparation",
            ],
        )

    with c4:
        participation = st.selectbox(
            "Participation and interaction",
            [
                "Not specified",
                "active and contributed positively to discussions and activities",
                "generally good but could be improved through more structured activities",
                "limited and should be strengthened in the next offering",
            ],
        )

        subjective_notes = st.text_area(
            "Additional subjective evaluation notes",
            placeholder="Optional: Add your own reflection on student engagement and learning experience.",
            height=90,
        )

    st.markdown("#### Course Improvement Insights")

    imp1 = st.checkbox("Add more real-world case studies", value=False)
    imp2 = st.checkbox("Expand hands-on modeling exercises", value=False)
    imp3 = st.checkbox("Explore AI-supported learning activities", value=False)

    improvement_notes = st.text_area(
        "Additional improvement actions",
        placeholder="Optional: Add additional improvements for the next course offering.",
        height=90,
    )

    if st.button("➕ Add Instructor Insight", type="primary"):
        with st.spinner("Combining instructor insights with AI-generated EOS text..."):
            crafted = craft_eos_text_with_ai(
                course=course,
                course_structure=course_structure,
                saved_results=saved_attainment_results,
                combined_so_attainment=combined_so_attainment,
                section_text=section_text,
                exam_feedback=exam_feedback,
                project_feedback=project_feedback,
                comparison_feedback=comparison_feedback,
                assessment_notes=assessment_notes,
                engagement=engagement,
                readiness=readiness,
                participation=participation,
                subjective_notes=subjective_notes,
                improvement_flags={
                    "case_studies": imp1,
                    "hands_on": imp2,
                    "ai_support": imp3,
                },
                improvement_notes=improvement_notes,
            )

            # Merge the newly crafted text into the final EOS sections.
            st.session_state["eos_assessment_based_evaluation"] = crafted.get("assessment_based_evaluation", "")
            st.session_state["eos_subjective_evaluation"] = crafted.get("subjective_evaluation", "")
            st.session_state["eos_course_improvement"] = crafted.get("course_improvement", "")

            st.success("Instructor insights were added and polished into the EOS draft.")

    # No separate final boxes for the three sections.
    # The final combined write-up appears only inside the editable EOS Draft.
    assessment_based_evaluation = st.session_state.get("eos_assessment_based_evaluation", "")
    subjective_evaluation = st.session_state.get("eos_subjective_evaluation", "")
    course_improvement = st.session_state.get("eos_course_improvement", "")

    if not assessment_based_evaluation or not subjective_evaluation or not course_improvement:
        fallback = craft_eos_text_with_ai(
            course=course,
            course_structure=course_structure,
            saved_results=saved_attainment_results,
            combined_so_attainment=combined_so_attainment,
            section_text=section_text,
            exam_feedback=exam_feedback,
            project_feedback=project_feedback,
            comparison_feedback=comparison_feedback,
            assessment_notes=assessment_notes,
            engagement=engagement,
            readiness=readiness,
            participation=participation,
            subjective_notes=subjective_notes,
            improvement_flags={
                "case_studies": imp1,
                "hands_on": imp2,
                "ai_support": imp3,
            },
            improvement_notes=improvement_notes,
        )

        assessment_based_evaluation = assessment_based_evaluation or fallback.get("assessment_based_evaluation", "")
        subjective_evaluation = subjective_evaluation or fallback.get("subjective_evaluation", "")
        course_improvement = course_improvement or fallback.get("course_improvement", "")

    draft_report = generate_eos_report(
        course={
            **course,
            "number_of_students": number_of_students,
        },
        course_structure=course_structure,
        combined_so_attainment=combined_so_attainment,
        assessment_based_evaluation=assessment_based_evaluation,
        subjective_evaluation=subjective_evaluation,
        course_improvement=course_improvement,
    )

    st.markdown("### 3. Editable EOS Draft")

    edited_eos_report = st.text_area(
        "EOS Draft",
        value=draft_report,
        height=650,
    )

    st.session_state["eos_report"] = edited_eos_report

    filename = f"EOS_Report_{course_code}_{academic_year}_{semester}_{today}.md"
    download_markdown("📥 Download EOS Report", st.session_state["eos_report"], filename)
