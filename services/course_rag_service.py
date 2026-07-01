import os
import re
import json
import zipfile
from collections import Counter

from services.ai_service import ask_ai_json


MATRIX_FOLDER = "articulation_matrix"
DOCUMENT_FOLDER = "course_documents"


# ==========================================================
# Public API
# ==========================================================

def build_course_rag_chunks(course, course_structure):
    """
    Build Course RAG v1 knowledge chunks.

    Sources:
    1. Structured course model
    2. Uploaded course documents
    3. Uploaded historical evidence

    This is intentionally lightweight:
    - no vector database
    - no external embeddings
    - keyword retrieval
    """

    chunks = []

    chunks.extend(_chunks_from_course_model(course, course_structure))
    chunks.extend(_chunks_from_uploaded_files(course))

    return chunks


def retrieve_relevant_chunks(query, chunks, top_k=8):
    if not query or not chunks:
        return []

    query_terms = _tokenize(query)

    scored = []

    for chunk in chunks:
        text = chunk.get("text", "")
        score = _score_text(query_terms, text)

        if score > 0:
            scored.append((score, chunk))

    scored.sort(key=lambda item: item[0], reverse=True)

    return [chunk for _, chunk in scored[:top_k]]


def answer_course_question(
    course,
    course_structure,
    query,
    conversation_history=None,
):
    chunks = build_course_rag_chunks(course, course_structure)
    retrieved = retrieve_relevant_chunks(query, chunks, top_k=10)

    context = "\n\n".join(
        f"[Source: {chunk.get('source', 'Course Knowledge')}]"
        f"\n{chunk.get('text', '')}"
        for chunk in retrieved
    )

    prompt = f"""
You are a Course Assistant for an ABET-oriented Curriculum Intelligence Platform.

Answer the instructor's question using ONLY the provided course context.
If the answer is not available in the context, say that the available course evidence does not contain enough information.

Be concise, practical, and instructor-friendly.

Course:
{course}

Course Structure:
{course_structure}

Retrieved Course Evidence:
{context}

Conversation History:
{conversation_history or []}

Instructor Question:
{query}

Return ONLY valid JSON in this format:
{{
    "answer": "string",
    "used_sources": ["string"],
    "follow_up_suggestion": "string"
}}
"""

    return ask_ai_json(prompt)


# ==========================================================
# Course Model Chunks
# ==========================================================

def _chunks_from_course_model(course, course_structure):
    chunks = []

    chunks.append(
        {
            "source": "Course Profile",
            "text": _safe_json(
                {
                    "course_code": course.get("course_code", ""),
                    "course_name": course.get("course_name", ""),
                    "department": course.get("department", ""),
                    "academic_year": course.get("academic_year", ""),
                    "semester": course.get("semester", ""),
                    "sections": course.get("sections", ""),
                    "instructor": course.get("instructor", ""),
                    "coordinator": course.get("coordinator", ""),
                }
            ),
        }
    )

    clos = course_structure.get("clos", [])
    if clos:
        chunks.append(
            {
                "source": "Course Learning Outcomes",
                "text": _safe_json(clos),
            }
        )

    student_outcomes = course_structure.get("student_outcomes", [])
    if student_outcomes:
        chunks.append(
            {
                "source": "Student Outcomes",
                "text": _safe_json(student_outcomes),
            }
        )

    assessments = course_structure.get(
        "assessment_instances",
        course_structure.get("assessments", []),
    )

    if assessments:
        chunks.append(
            {
                "source": "Assessment Structure",
                "text": _safe_json(assessments),
            }
        )

        for assessment in assessments:
            chunks.append(
                {
                    "source": f"Assessment: {assessment.get('name', '')}",
                    "text": _safe_json(assessment),
                }
            )

    stats = course_structure.get("statistics", {})
    if stats:
        chunks.append(
            {
                "source": "Course Statistics",
                "text": _safe_json(stats),
            }
        )

    # Other generated AI outputs
    extra_state_keys = [
        ("Curriculum Analysis", "curriculum_analysis"),
        ("Course Enhancement Plan", "course_enhancement_plan"),
        ("SO Attainment", "so_attainment"),
        ("SO Attainment Insights", "so_attainment_ai"),
    ]

    try:
        import streamlit as st

        for source_name, key in extra_state_keys:
            value = st.session_state.get(key)
            if value:
                chunks.append(
                    {
                        "source": source_name,
                        "text": _safe_json(value),
                    }
                )
    except Exception:
        pass

    return chunks


# ==========================================================
# Uploaded File Chunks
# ==========================================================

def _chunks_from_uploaded_files(course):
    chunks = []

    file_fields = {
        "matrix_file": MATRIX_FOLDER,
        "syllabus_file": DOCUMENT_FOLDER,
        "lab_syllabus_file": DOCUMENT_FOLDER,
        "slides_file": DOCUMENT_FOLDER,
        "labs_file": DOCUMENT_FOLDER,
        "assessments_file": DOCUMENT_FOLDER,
        "esr_file": DOCUMENT_FOLDER,
        "coordination_file": DOCUMENT_FOLDER,
        "rubrics_file": DOCUMENT_FOLDER,
    }

    for field, folder in file_fields.items():
        for filename in _normalize_file_list(course.get(field)):
            path = os.path.join(folder, filename)

            if not filename or not os.path.exists(path):
                continue

            text = _read_file_text(path)

            if not text:
                continue

            for index, part in enumerate(_split_text(text, chunk_size=1800), start=1):
                chunks.append(
                    {
                        "source": f"{field}: {filename} (part {index})",
                        "text": part,
                    }
                )

    return chunks


def _normalize_file_list(value):
    if not value:
        return []

    if isinstance(value, list):
        return value

    return [value]


def _read_file_text(path):
    ext = os.path.splitext(path)[1].lower()

    try:
        if ext in [".txt", ".md", ".csv"]:
            return _read_text_file(path)

        if ext in [".docx"]:
            return _read_docx_file(path)

        if ext in [".pdf"]:
            return _read_pdf_file(path)

        if ext in [".xlsx", ".xlsm", ".xls"]:
            return _read_excel_file(path)

        if ext in [".zip"]:
            return _read_zip_file(path)

        if ext in [".py", ".ipynb", ".json"]:
            return _read_text_file(path)

    except Exception as e:
        return f"Could not read file {os.path.basename(path)}: {e}"

    return ""


def _read_text_file(path):
    for encoding in ["utf-8", "latin-1"]:
        try:
            with open(path, "r", encoding=encoding) as f:
                return f.read()
        except Exception:
            continue

    return ""


def _read_docx_file(path):
    try:
        from docx import Document

        doc = Document(path)
        return "\n".join(p.text for p in doc.paragraphs if p.text.strip())
    except Exception:
        return ""


def _read_pdf_file(path):
    try:
        import PyPDF2

        text = []
        with open(path, "rb") as f:
            reader = PyPDF2.PdfReader(f)

            for page in reader.pages:
                page_text = page.extract_text() or ""
                text.append(page_text)

        return "\n".join(text)
    except Exception:
        try:
            from pypdf import PdfReader

            text = []
            reader = PdfReader(path)

            for page in reader.pages:
                text.append(page.extract_text() or "")

            return "\n".join(text)
        except Exception:
            return ""


def _read_excel_file(path):
    try:
        import openpyxl

        workbook = openpyxl.load_workbook(path, data_only=True)
        lines = []

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            lines.append(f"Sheet: {sheet_name}")

            for row in ws.iter_rows(values_only=True):
                values = [str(v) for v in row if v is not None]
                if values:
                    lines.append(" | ".join(values))

        return "\n".join(lines)
    except Exception:
        return ""


def _read_zip_file(path):
    lines = []

    try:
        with zipfile.ZipFile(path, "r") as z:
            for name in z.namelist():
                ext = os.path.splitext(name)[1].lower()

                if ext not in [".txt", ".md", ".csv", ".py", ".json"]:
                    lines.append(f"File in archive: {name}")
                    continue

                try:
                    content = z.read(name).decode("utf-8", errors="ignore")
                    lines.append(f"File in archive: {name}\n{content[:3000]}")
                except Exception:
                    lines.append(f"File in archive: {name}")

        return "\n\n".join(lines)

    except Exception:
        return ""


# ==========================================================
# Retrieval Helpers
# ==========================================================

def _split_text(text, chunk_size=1800):
    text = re.sub(r"\s+", " ", str(text)).strip()

    if not text:
        return []

    chunks = []

    for start in range(0, len(text), chunk_size):
        chunks.append(text[start:start + chunk_size])

    return chunks


def _tokenize(text):
    tokens = re.findall(r"[A-Za-z0-9]+", str(text).lower())
    stopwords = {
        "the", "and", "or", "of", "to", "a", "in", "for", "on", "with",
        "is", "are", "this", "that", "what", "which", "how", "does",
        "do", "me", "my", "course", "assessment",
    }

    return [t for t in tokens if t not in stopwords and len(t) > 1]


def _score_text(query_terms, text):
    if not query_terms:
        return 0

    text_terms = Counter(_tokenize(text))
    score = 0

    for term in query_terms:
        score += text_terms.get(term, 0)

    # Boost exact phrase-ish matches
    lower_text = str(text).lower()
    for term in query_terms:
        if term in lower_text:
            score += 1

    return score


def _safe_json(value):
    try:
        return json.dumps(value, ensure_ascii=False, indent=2)
    except Exception:
        return str(value)
