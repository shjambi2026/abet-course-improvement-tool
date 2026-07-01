import openpyxl


PASS_THRESHOLD_PERCENT = 65
SO_TARGET_PERCENT = 65


def calculate_so_attainment_from_raw_workbook(uploaded_workbook):
    """
    Calculate SO attainment from the uploaded SO marks workbook.

    Expected structure:
        Program | Section | Student ID | SO2 | SO3 | ...
        Maximum Marks | | | 30 | 20
        IS | DAR | 123456 | 25 | 18

    Student Name is not required.
    """

    wb = openpyxl.load_workbook(uploaded_workbook, data_only=True)
    ws = wb["grades"] if "grades" in wb.sheetnames else wb.active

    header_row = _find_header_row(ws)

    if header_row is None:
        raise ValueError("Could not find the marks table header row.")

    headers = [
        ws.cell(row=header_row, column=col).value
        for col in range(1, ws.max_column + 1)
    ]

    so_columns = {}

    for col, header in enumerate(headers, start=1):
        text = str(header or "").strip().upper().replace(" ", "")

        if text.startswith("SO") and text[2:].isdigit():
            so_columns[text] = col

    if not so_columns:
        raise ValueError("No SO mark columns were found. Expected columns such as SO2 or SO3.")

    max_marks_row = header_row + 1

    first_student_row = header_row + 2

    summary = []

    for so, col in sorted(so_columns.items(), key=lambda item: int(item[0].replace("SO", ""))):
        max_marks = ws.cell(row=max_marks_row, column=col).value

        if not isinstance(max_marks, (int, float)) or max_marks <= 0:
            raise ValueError(
                f"Maximum marks for {so} is missing or invalid. "
                "Please enter maximum marks in the row directly under the header."
            )

        student_marks = []

        for row in range(first_student_row, ws.max_row + 1):
            student_id = ws.cell(row=row, column=3).value
            mark = ws.cell(row=row, column=col).value

            if student_id is None and mark is None:
                continue

            if mark is None:
                continue

            if isinstance(mark, (int, float)):
                student_marks.append(float(mark))

        students_assessed = len(student_marks)

        pass_mark = max_marks * (PASS_THRESHOLD_PERCENT / 100)
        students_attained = sum(1 for mark in student_marks if mark >= pass_mark)

        attainment_percent = (
            round((students_attained / students_assessed) * 100, 2)
            if students_assessed
            else 0
        )

        summary.append(
            {
                "so": so,
                "maximum_marks": max_marks,
                "pass_mark": round(pass_mark, 2),
                "students_assessed": students_assessed,
                "students_attained": students_attained,
                "attainment_percent": attainment_percent,
                "target_percent": SO_TARGET_PERCENT,
                "target_met": "Yes" if attainment_percent >= SO_TARGET_PERCENT else "No",
            }
        )

    statistics = {
        "assessed_sos": len(summary),
        "sos_met_target": sum(1 for item in summary if item["target_met"] == "Yes"),
        "average_attainment": round(
            sum(item["attainment_percent"] for item in summary) / len(summary),
            2,
        )
        if summary
        else 0,
    }

    return {
        "summary": summary,
        "statistics": statistics,
        "source": "Uploaded SO marks workbook",
    }


def build_so_attainment_report(course, attainment, ai=None):
    lines = []

    lines.append("# Student Outcome Attainment Report")
    lines.append("")
    lines.append(f"**Course:** {course.get('course_code', '')} – {course.get('course_name', '')}")
    lines.append(f"**Semester:** {course.get('semester', '')}")
    lines.append(f"**Academic Year:** {course.get('academic_year', '')}")
    lines.append("")

    lines.append("## Attainment Summary")
    lines.append("")

    for item in attainment.get("summary", []):
        lines.append(
            f"- **{item.get('so', '')}:** "
            f"{item.get('attainment_percent', 0)}% "
            f"({item.get('students_attained', 0)}/{item.get('students_assessed', 0)} students attained; "
            f"Target met: {item.get('target_met', 'No')})"
        )

    if ai:
        lines.append("")
        lines.append("## Assessment-Based Evaluation")
        lines.append(ai.get("assessment_based_evaluation", ""))

        lines.append("")
        lines.append("## Recommended Improvement Actions")

        for item in ai.get("improvement_actions", []):
            lines.append(f"- {item}")

        lines.append("")
        lines.append("## Overall AI Observation")
        lines.append(ai.get("overall_observation", ""))

    return "\n".join(lines)


def _find_header_row(ws):
    for row in range(1, ws.max_row + 1):
        values = [
            str(ws.cell(row=row, column=col).value or "").strip().lower()
            for col in range(1, min(ws.max_column, 8) + 1)
        ]

        if "program" in values and "section" in values and "student id" in values:
            return row

    return None
