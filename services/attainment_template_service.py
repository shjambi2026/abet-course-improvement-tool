from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


HEADER_FILL = PatternFill(
    fill_type="solid",
    start_color="1F7A4C",
    end_color="1F7A4C",
)

HEADER_FONT = Font(
    bold=True,
    color="FFFFFF",
)

BOLD = Font(bold=True)

PROGRAM_OPTIONS = ["IS", "CS", "IT"]


def _split_sections(value):
    if not value:
        return [""]

    text = str(value)

    for sep in [",", "/", ";", "|"]:
        text = text.replace(sep, ",")

    sections = [s.strip() for s in text.split(",") if s.strip()]

    return sections or [""]


def _sort_so_ids(values):
    def key(value):
        try:
            return int(str(value).replace("SO", ""))
        except Exception:
            return 9999

    return sorted(values or [], key=key)


def _get_assessed_sos(assessment=None, assessed_sos=None):
    if assessment:
        distribution = assessment.get("so_distribution", {})
        if distribution:
            return _sort_so_ids(distribution.keys())

        sos = assessment.get("assessed_sos", assessment.get("sos", []))
        return _sort_so_ids(sos)

    return _sort_so_ids(assessed_sos or [])


def generate_so_attainment_template(
    course,
    assessment=None,
    assessment_name=None,
    assessed_sos=None,
):
    """
    Generate an SO marks workbook.

    Required upload structure:
        Program | Section | Student ID | SO2 | SO3 | ...

    The first row after the header is:
        Maximum Marks | | | <max for SO2> | <max for SO3>

    Student rows start after the Maximum Marks row.
    """

    assessed_sos = _get_assessed_sos(
        assessment=assessment,
        assessed_sos=assessed_sos,
    )

    if assessment:
        assessment_name = assessment.get("name", assessment_name or "Assessment")
    else:
        assessment_name = assessment_name or "Assessment"

    wb = Workbook()
    ws = wb.active
    ws.title = "grades"

    # ======================================================
    # Course and Assessment Information
    # ======================================================

    info = [
        ("Course Code", course.get("course_code", "")),
        ("Course Name", course.get("course_name", "")),
        ("Academic Year", course.get("academic_year", "")),
        ("Semester", course.get("semester", "")),
        ("Instructor", course.get("instructor", "")),
        ("Course Section(s)", course.get("sections", course.get("section", ""))),
        ("Assessment", assessment_name),
    ]

    if assessment:
        info.extend(
            [
                ("Assessment Week", assessment.get("week", "")),
                ("Assessment Type", assessment.get("type", "")),
                ("Weight (%)", assessment.get("course_weight", assessment.get("weight", ""))),
            ]
        )

    for row, (label, value) in enumerate(info, start=1):
        ws.cell(row=row, column=1).value = label
        ws.cell(row=row, column=2).value = value
        ws.cell(row=row, column=1).font = BOLD

    # ======================================================
    # Marks Table
    # ======================================================

    start_row = 13

    headers = [
        "Program",
        "Section",
        "Student ID",
    ] + assessed_sos

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col)
        cell.value = header
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    max_row = start_row + 1
    ws.cell(row=max_row, column=1).value = "Maximum Marks"
    ws.cell(row=max_row, column=1).font = BOLD

    first_student_row = start_row + 2
    last_student_row = start_row + 101

    # Program and Section cells are intentionally left blank.
    # Users must select valid values from dropdown lists.

    # ======================================================
    # Data Validation
    # ======================================================

    program_validation = DataValidation(
        type="list",
        formula1='"IS,CS,IT"',
        allow_blank=False,
    )
    ws.add_data_validation(program_validation)
    program_validation.add(f"A{first_student_row}:A{last_student_row}")

    sections = _split_sections(course.get("sections", course.get("section", "")))

    if sections and sections != [""]:
        section_formula = '"' + ",".join(sections) + '"'
        section_validation = DataValidation(
            type="list",
            formula1=section_formula,
            allow_blank=False,
        )
        ws.add_data_validation(section_validation)
        section_validation.add(f"B{first_student_row}:B{last_student_row}")

    # ======================================================
    # Formatting
    # ======================================================

    widths = {
        "A": 14,
        "B": 16,
        "C": 18,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    first_so_column = 4

    for i in range(len(assessed_sos)):
        col_letter = ws.cell(row=start_row, column=first_so_column + i).column_letter
        ws.column_dimensions[col_letter].width = 14

    ws.freeze_panes = f"A{first_student_row}"

    # ======================================================
    # Instructions Sheet
    # ======================================================

    inst = wb.create_sheet("Instructions")

    inst["A1"] = "Student Outcome Marks Workbook"
    inst["A1"].font = Font(size=14, bold=True)

    instructions = [
        "",
        "Instructions",
        "",
        "1. Do not rename the 'grades' worksheet.",
        "2. Do not modify the header row.",
        "3. Program must be selected from the dropdown list: IS, CS, IT.",
        "4. Section must be selected from the course section dropdown list.",
        "5. Enter the maximum marks for each assessed Student Outcome in the 'Maximum Marks' row.",
        "6. Enter one row per student.",
        "7. Student Name is intentionally excluded; Student ID is sufficient.",
        "8. Record only the marks corresponding to the assessed Student Outcomes.",
        "9. Upload the completed workbook to the Student Outcome Attainment page.",
    ]

    for row, text in enumerate(instructions, start=2):
        inst.cell(row=row, column=1).value = text

    inst.column_dimensions["A"].width = 110

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return output
